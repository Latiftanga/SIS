# students/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.http import HttpRequest, HttpResponse
from django.db.models import Q
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
from .models import Student, Programme
from .forms import StudentCreateForm, StudentBulkImportForm, ProgrammeForm
from accounts.models import User
from accounts.decorators import school_admin_required
from accounts.utils import generate_secure_password
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from io import BytesIO


@school_admin_required
def student_list(request: HttpRequest) -> HttpResponse:
    """List all students with HTMX support"""
    students = Student.objects.select_related('user').all()

    # Status filter
    status = request.GET.get('status', '')
    if status == 'active':
        students = students.filter(is_active=True)
    elif status == 'inactive':
        students = students.filter(is_active=False)

    # Gender filter
    gender = request.GET.get('gender', '')
    if gender:
        students = students.filter(gender=gender)

    # Class filter (grade is determined by class enrollment)
    class_filter = request.GET.get('class', '')
    if class_filter:
        from classes.models import StudentEnrollment
        # Get students enrolled in the specified class
        enrolled_students = StudentEnrollment.objects.filter(
            class_obj__name=class_filter,
            is_active=True
        ).values_list('student_id', flat=True)
        students = students.filter(id__in=enrolled_students)

    # Search functionality
    search = request.GET.get('search', '')
    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(student_id__icontains=search) |
            Q(guardian_name__icontains=search)
        )

    # Get unique classes for filter dropdown
    from classes.models import Class
    classes = Class.objects.filter(is_active=True).order_by('grade_level', 'section')

    context = {
        'students': students,
        'search': search,
        'status': status,
        'gender': gender,
        'class_filter': class_filter,
        'classes': classes,
    }

    # Return partial template for HTMX requests
    if request.headers.get('HX-Request'):
        return render(request, 'students/partials/student_list_content.html', context)

    # Full page render
    return render(request, 'students/student_list.html', context)


@school_admin_required
def student_create(request: HttpRequest) -> HttpResponse:
    """Show create student form or handle form submission"""
    if request.method == 'POST':
        form = StudentCreateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Save with request for email URL building
                student, generated_password = form.save(request=request)

                # Build success message
                success_message = f'Student {student.get_full_name()} added successfully!'
                if generated_password:
                    email = student.email or student.guardian_email
                    success_message += f' Credentials sent to {email}'

                messages.success(request, success_message)
                return redirect('students:list')
            except Exception as e:
                messages.error(request, f'Error creating student: {str(e)}')
    else:
        # GET request - show empty form
        form = StudentCreateForm()

    context = {
        'form': form,
        'action': 'Add'
    }

    # Return partial template for HTMX requests
    if request.headers.get('HX-Request'):
        return render(request, 'students/partials/student_form_content.html', context)

    # Full page render
    return render(request, 'students/student_create.html', context)


@school_admin_required
def student_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """View student details"""
    student = get_object_or_404(Student, pk=pk)

    # For HTMX requests, return inline detail content
    if request.headers.get('HX-Request'):
        response = render(request, 'students/partials/student_detail_inline.html', {'student': student})
        response['HX-Trigger'] = 'formLoaded'
        return response

    return render(request, 'students/student_detail_page.html', {
        'student': student
    })


@school_admin_required
def student_detail_pdf(request: HttpRequest, pk: int) -> HttpResponse:
    """Generate and download PDF of student details"""
    student = get_object_or_404(Student, pk=pk)

    # Render the HTML template for PDF
    html_string = render_to_string('students/student_detail_pdf.html', {
        'student': student
    }, request=request)

    # Create PDF
    font_config = FontConfiguration()
    html = HTML(string=html_string, base_url=request.build_absolute_uri())

    # Generate PDF with custom CSS
    pdf_css = CSS(string='''
        @page {
            size: A4;
            margin: 2cm;
        }
        body {
            font-family: 'DejaVu Sans', Arial, sans-serif;
            font-size: 11pt;
            line-height: 1.6;
            color: #333;
        }
        h1 {
            color: #2563eb;
            font-size: 24pt;
            margin-bottom: 10pt;
            border-bottom: 2pt solid #2563eb;
            padding-bottom: 5pt;
        }
        h2 {
            color: #1e40af;
            font-size: 16pt;
            margin-top: 15pt;
            margin-bottom: 8pt;
            border-bottom: 1pt solid #ddd;
            padding-bottom: 3pt;
        }
        .info-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 10pt;
            margin-bottom: 15pt;
        }
        .info-item {
            margin-bottom: 8pt;
        }
        .info-label {
            font-weight: bold;
            color: #666;
            font-size: 9pt;
            text-transform: uppercase;
        }
        .info-value {
            font-size: 11pt;
            color: #333;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10pt;
        }
        th {
            background-color: #2563eb;
            color: white;
            padding: 8pt;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 6pt 8pt;
            border-bottom: 1pt solid #ddd;
        }
        tr:nth-child(even) {
            background-color: #f9fafb;
        }
        .header-section {
            margin-bottom: 20pt;
            padding-bottom: 15pt;
            border-bottom: 2pt solid #e5e7eb;
        }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 10pt;
            margin-bottom: 20pt;
        }
        .stat-card {
            background: #f3f4f6;
            padding: 10pt;
            border-left: 3pt solid #2563eb;
        }
        .stat-label {
            font-size: 9pt;
            color: #666;
            text-transform: uppercase;
        }
        .stat-value {
            font-size: 16pt;
            font-weight: bold;
            color: #1e40af;
        }
    ''', font_config=font_config)

    pdf = html.write_pdf(stylesheets=[pdf_css], font_config=font_config)

    # Create response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_{student.student_id}_{student.get_full_name().replace(" ", "_")}.pdf"'

    return response


@school_admin_required
def student_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit student information"""
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        form = StudentCreateForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            try:
                # Save changes
                student, _ = form.save(request=request)

                messages.success(request, f'Student {student.get_full_name()} updated successfully!')
                return redirect('students:list')
            except Exception as e:
                messages.error(request, f'Error updating student: {str(e)}')
    else:
        # GET request - show form with student data
        form = StudentCreateForm(instance=student)

    context = {
        'form': form,
        'action': 'Edit',
        'student': student
    }

    # Return partial template for HTMX requests
    if request.headers.get('HX-Request'):
        return render(request, 'students/partials/student_form_content.html', context)

    # Full page render
    return render(request, 'students/student_create.html', context)


@school_admin_required
@require_http_methods(["POST", "DELETE"])
def student_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Soft delete student"""
    student = get_object_or_404(Student, pk=pk)
    student.is_active = False
    student.save()

    # Return empty response to trigger refresh
    return HttpResponse(headers={'HX-Trigger': 'studentDeleted'})


@school_admin_required
@require_http_methods(["GET", "POST"])
def student_bulk_import(request: HttpRequest) -> HttpResponse:
    """
    Bulk import students from Excel or CSV file.
    Shows upload form and preview of data to be imported.
    """
    if request.method == 'POST':
        form = StudentBulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Parse file
                students_data = form.parse_file()
                create_accounts = form.cleaned_data['create_user_accounts']

                # Store data in session for processing
                request.session['bulk_import_data'] = students_data
                request.session['bulk_import_create_accounts'] = create_accounts

                # Calculate statistics
                valid_count = sum(1 for s in students_data if s['valid'])
                invalid_count = len(students_data) - valid_count

                context = {
                    'students': students_data,
                    'total': len(students_data),
                    'valid': valid_count,
                    'invalid': invalid_count,
                    'create_accounts': create_accounts,
                }

                # For HTMX requests, return preview content partial
                if request.headers.get('HX-Request'):
                    response = render(request, 'students/partials/bulk_import_preview_content.html', context)
                    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                    response['Pragma'] = 'no-cache'
                    response['Expires'] = '0'
                    return response

                return render(request, 'students/bulk_import_preview.html', context)

            except Exception as e:
                messages.error(request, f'Error parsing file: {str(e)}')
                # For HTMX requests, return form with error
                if request.headers.get('HX-Request'):
                    response = render(request, 'students/partials/bulk_import_content.html', {'form': form})
                    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                    response['Pragma'] = 'no-cache'
                    response['Expires'] = '0'
                    return response
    else:
        form = StudentBulkImportForm()

    # For HTMX requests, return content partial
    if request.headers.get('HX-Request'):
        response = render(request, 'students/partials/bulk_import_content.html', {'form': form})
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    return render(request, 'students/bulk_import.html', {'form': form})


@school_admin_required
@require_http_methods(["POST"])
def student_bulk_import_process(request: HttpRequest) -> HttpResponse:
    """
    Process bulk import after preview confirmation.
    Creates student records and optionally user accounts.
    """
    students_data = request.session.get('bulk_import_data', [])
    create_accounts = request.session.get('bulk_import_create_accounts', False)

    if not students_data:
        messages.error(request, 'No import data found. Please upload a file first.')
        return redirect('students:bulk_import')

    success_count = 0
    error_count = 0
    errors = []

    for student_data in students_data:
        # Skip invalid rows
        if not student_data['valid']:
            error_count += 1
            continue

        try:
            # Create student
            student = Student(
                first_name=student_data['first_name'],
                last_name=student_data['last_name'],
                other_names=student_data.get('other_names', ''),
                email=student_data.get('email', ''),
                gender=student_data['gender'],
                phone_number=student_data.get('phone_number', ''),
                residential_address=student_data.get('residential_address', ''),
                student_id=student_data['student_id'],
                admission_date=student_data['admission_date'],
                date_of_birth=student_data['date_of_birth'],
                guardian_name=student_data['guardian_name'],
                guardian_relationship=student_data.get('guardian_relationship', 'Parent'),
                guardian_phone=student_data['guardian_phone'],
                guardian_email=student_data.get('guardian_email', ''),
                guardian_address=student_data.get('guardian_address', ''),
                emergency_contact_name=student_data.get('emergency_contact_name', ''),
                emergency_contact_phone=student_data.get('emergency_contact_phone', ''),
                medical_conditions=student_data.get('medical_conditions', ''),
            )

            # Create user account if requested and student has email
            should_create_account = (create_accounts or student_data.get('create_account', False)) and student_data.get('email')
            if should_create_account:
                # Generate password
                password = generate_secure_password()

                # Create user
                user = User.objects.create_student(
                    email=student_data['email'],
                    password=password
                )
                user.force_password_change = True
                user.save()

                student.user = user

            student.save()

            # Create class enrollment if class_name is provided
            if student_data.get('class_name'):
                from classes.models import Class, StudentEnrollment

                try:
                    class_obj = Class.objects.get(name=student_data['class_name'], is_active=True)
                    # Use the class's academic year
                    academic_year = class_obj.academic_year if class_obj.academic_year else '2024/2025'

                    # Create enrollment
                    StudentEnrollment.objects.create(
                        student=student,
                        class_obj=class_obj,
                        academic_year=academic_year,
                        enrollment_date=student.admission_date,
                        status='enrolled',
                        is_active=True
                    )
                except Class.DoesNotExist:
                    # Class validation should have caught this, but just in case
                    pass

            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {student_data['row_num']}: {str(e)}")

    # Clear session data
    if 'bulk_import_data' in request.session:
        del request.session['bulk_import_data']
    if 'bulk_import_create_accounts' in request.session:
        del request.session['bulk_import_create_accounts']

    # Show results
    if success_count > 0:
        messages.success(
            request,
            f'Successfully imported {success_count} student(s).'
        )

    if error_count > 0:
        error_msg = f'Failed to import {error_count} student(s).'
        if errors:
            error_msg += ' Errors: ' + '; '.join(errors[:5])
        messages.error(request, error_msg)

    # For HTMX requests, redirect to student list
    if request.headers.get('HX-Request'):
        response = HttpResponse()
        response['HX-Redirect'] = '/students/'
        return response

    return redirect('students:list')


@school_admin_required
def student_download_template(request: HttpRequest) -> HttpResponse:
    """
    Download sample Excel template for bulk import.
    """
    file_format = request.GET.get('format', 'xlsx')

    if file_format == 'csv':
        return _generate_csv_template()
    else:
        return _generate_excel_template()


def _generate_excel_template() -> HttpResponse:
    """Generate Excel template with sample data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Define headers
    headers = [
        'first_name', 'last_name', 'other_names', 'date_of_birth',
        'gender', 'email', 'phone_number', 'residential_address',
        'student_id', 'admission_date', 'class_name',
        'guardian_name', 'guardian_relationship', 'guardian_phone',
        'guardian_email', 'guardian_address',
        'emergency_contact_name', 'emergency_contact_phone',
        'medical_conditions', 'create_account'
    ]

    # Style headers
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Get existing active classes for sample data
    from classes.models import Class
    all_active_classes = Class.objects.filter(is_active=True).order_by('grade_level', 'section')

    # Use real class names if available, otherwise use placeholders
    class_1 = all_active_classes[0].name if all_active_classes else '[REPLACE_WITH_ACTUAL_CLASS_NAME]'
    class_2 = all_active_classes[1].name if len(all_active_classes) > 1 else '[REPLACE_WITH_ACTUAL_CLASS_NAME]'

    # Add sample data
    sample_data = [
        [
            'Kwame', 'Mensah', 'Kofi', '2010-03-15',
            'Male', 'kwame.mensah@example.com', '+233244123456', '123 Main St, Accra',
            'STU001', '2024-09-01', class_1,
            'Abena Mensah', 'Mother', '+233244987654',
            'abena.mensah@example.com', '123 Main St, Accra',
            'Yaw Mensah', '+233244111222',
            '', 'yes'
        ],
        [
            'Akua', 'Owusu', '', '2012-07-20',
            'Female', '', '0244567890', '456 Oak Ave, Kumasi',
            'STU002', '2024-09-01', class_2,
            'Kwesi Owusu', 'Father', '0244333444',
            'kwesi.owusu@example.com', '456 Oak Ave, Kumasi',
            '', '',
            'Asthma - requires inhaler', 'no'
        ],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add a second sheet with available classes
    ws2 = wb.create_sheet("Available Classes")
    ws2['A1'] = "Available Class Names (Copy Exact Names)"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].fill = PatternFill(start_color="34D399", end_color="34D399", fill_type="solid")

    ws2['A2'] = "IMPORTANT: Use exact class names from this list in the class_name column"
    ws2['A2'].font = Font(italic=True, color="DC2626")

    # List all active classes
    row_num = 4
    ws2[f'A{row_num}'] = "Class Name"
    ws2[f'A{row_num}'].font = Font(bold=True)
    row_num += 1

    for class_obj in all_active_classes:
        ws2[f'A{row_num}'] = class_obj.name
        row_num += 1

    if not all_active_classes:
        ws2['A5'] = "No active classes found. Please create classes first."
        ws2['A5'].font = Font(color="DC2626")

    ws2.column_dimensions['A'].width = 40

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="students_import_template_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response


def _generate_csv_template() -> HttpResponse:
    """Generate CSV template with sample data."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="students_import_template_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)

    # Get all active classes
    from classes.models import Class
    all_active_classes = Class.objects.filter(is_active=True).order_by('grade_level', 'section')

    # Write comments with available classes
    writer.writerow(['# IMPORTANT: Use exact class names from the list below in the class_name column'])
    writer.writerow(['# Available Classes: ' + ', '.join([c.name for c in all_active_classes]) if all_active_classes else '# No active classes found - please create classes first'])
    writer.writerow([])  # Empty row for spacing

    # Write headers
    headers = [
        'first_name', 'last_name', 'other_names', 'date_of_birth',
        'gender', 'email', 'phone_number', 'residential_address',
        'student_id', 'admission_date', 'class_name',
        'guardian_name', 'guardian_relationship', 'guardian_phone',
        'guardian_email', 'guardian_address',
        'emergency_contact_name', 'emergency_contact_phone',
        'medical_conditions', 'create_account'
    ]
    writer.writerow(headers)

    # Get existing active classes for sample data
    from classes.models import Class
    all_active_classes = Class.objects.filter(is_active=True).order_by('grade_level', 'section')

    # Use real class names if available, otherwise use placeholders
    class_1 = all_active_classes[0].name if all_active_classes else '[REPLACE_WITH_ACTUAL_CLASS_NAME]'
    class_2 = all_active_classes[1].name if len(all_active_classes) > 1 else '[REPLACE_WITH_ACTUAL_CLASS_NAME]'

    # Write sample data
    sample_data = [
        [
            'Kwame', 'Mensah', 'Kofi', '2010-03-15',
            'Male', 'kwame.mensah@example.com', '+233244123456', '123 Main St, Accra',
            'STU001', '2024-09-01', class_1,
            'Abena Mensah', 'Mother', '+233244987654',
            'abena.mensah@example.com', '123 Main St, Accra',
            'Yaw Mensah', '+233244111222',
            '', 'yes'
        ],
        [
            'Akua', 'Owusu', '', '2012-07-20',
            'Female', '', '0244567890', '456 Oak Ave, Kumasi',
            'STU002', '2024-09-01', class_2,
            'Kwesi Owusu', 'Father', '0244333444',
            'kwesi.owusu@example.com', '456 Oak Ave, Kumasi',
            '', '',
            'Asthma - requires inhaler', 'no'
        ],
    ]

    for row in sample_data:
        writer.writerow(row)

    return response


# ==================== Programme Views ====================


@school_admin_required
def programme_list(request: HttpRequest) -> HttpResponse:
    """List all programmes with HTMX support"""
    programmes = Programme.objects.all()

    # Status filter
    status = request.GET.get('status', '')
    if status == 'active':
        programmes = programmes.filter(is_active=True)
    elif status == 'inactive':
        programmes = programmes.filter(is_active=False)
    else:
        # Default to showing only active programmes
        programmes = programmes.filter(is_active=True)

    # Search functionality
    search = request.GET.get('search', '')
    if search:
        programmes = programmes.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )

    context = {
        'programmes': programmes,
        'search': search,
        'status': status,
    }

    # If HTMX request for search/filters, return only the table rows
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') == 'programmes-table':
        return render(request, 'students/programmes/partials/programme_rows.html', context)

    # If HTMX request for navigation, return full content partial
    if request.headers.get('HX-Request'):
        return render(request, 'students/programmes/partials/programme_list_content.html', context)

    # Full page render
    return render(request, 'students/programmes/programme_list.html', context)


@school_admin_required
def programme_create(request: HttpRequest) -> HttpResponse:
    """Show create programme form or handle form submission"""
    if request.method == 'POST':
        form = ProgrammeForm(request.POST)
        if form.is_valid():
            try:
                programme = form.save()
                messages.success(request, f'Programme "{programme.name}" created successfully!')
                return redirect('students:programme_list')
            except Exception as e:
                messages.error(request, f'Error creating programme: {str(e)}')
    else:
        form = ProgrammeForm()

    context = {
        'form': form,
        'action': 'Create'
    }

    # Return partial template for HTMX requests
    if request.headers.get('HX-Request'):
        return render(request, 'students/programmes/partials/programme_form_content.html', context)

    return render(request, 'students/programmes/programme_form.html', context)


@school_admin_required
def programme_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit existing programme"""
    programme = get_object_or_404(Programme, pk=pk)

    if request.method == 'POST':
        form = ProgrammeForm(request.POST, instance=programme)
        if form.is_valid():
            try:
                programme = form.save()
                messages.success(request, f'Programme "{programme.name}" updated successfully!')
                return redirect('students:programme_list')
            except Exception as e:
                messages.error(request, f'Error updating programme: {str(e)}')
    else:
        form = ProgrammeForm(instance=programme)

    context = {
        'form': form,
        'action': 'Edit',
        'programme': programme
    }

    # Return partial template for HTMX requests
    if request.headers.get('HX-Request'):
        return render(request, 'students/programmes/partials/programme_form_content.html', context)

    return render(request, 'students/programmes/programme_form.html', context)


@school_admin_required
@require_http_methods(["DELETE"])
def programme_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Soft delete programme"""
    programme = get_object_or_404(Programme, pk=pk)
    programme.is_active = False
    programme.save()

    # Return empty response to remove the row
    return HttpResponse(headers={'HX-Trigger': 'programmeDeleted'})
