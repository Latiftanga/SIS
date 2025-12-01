# teachers/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.http import HttpRequest, HttpResponse
from django.db.models import Q
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from .models import Teacher
from .forms import TeacherCreateForm, TeacherBulkImportForm
from accounts.models import User
from accounts.decorators import school_admin_required
from accounts.utils import generate_secure_password
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from io import BytesIO
from weasyprint import HTML


@school_admin_required
def teacher_list(request: HttpRequest) -> HttpResponse:
    """List all teachers with HTMX support and pagination"""
    teachers = Teacher.objects.select_related('user').all()

    # Status filter
    status = request.GET.get('status', '')
    if status == 'active':
        teachers = teachers.filter(is_active=True)
    elif status == 'inactive':
        teachers = teachers.filter(is_active=False)
    else:
        # Default to showing only active teachers if no filter specified
        teachers = teachers.filter(is_active=True)

    # Gender filter
    gender = request.GET.get('gender', '')
    if gender:
        teachers = teachers.filter(gender=gender)

    # Search functionality
    search = request.GET.get('search', '')
    if search:
        teachers = teachers.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(employee_id__icontains=search)
        )

    # Pagination - 20 teachers per page (modern load more approach)
    page_number = request.GET.get('page', 1)
    paginator = Paginator(teachers, 20)  # 20 items per page
    page_obj = paginator.get_page(page_number)

    # Count teachers with accounts on current page
    teachers_with_accounts = sum(1 for teacher in page_obj if teacher.user is not None)

    context = {
        'teachers': page_obj,
        'page_obj': page_obj,
        'teachers_with_accounts': teachers_with_accounts,
        'search': search,
        'status': status,
        'gender': gender,
    }

    # If HTMX request for "load more", return only the new rows + load more button
    if request.headers.get('HX-Request') and request.GET.get('page'):
        return render(request, 'teachers/partials/teacher_rows_paginated.html', context)

    # If HTMX request for search/filters, return table with pagination reset
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') == 'teachers-table':
        return render(request, 'teachers/partials/teacher_rows.html', context)

    # Full page render (works for both initial load and HTMX navigation)
    return render(request, 'teachers/teacher_list.html', context)


@school_admin_required
def teacher_create(request: HttpRequest) -> HttpResponse:
    """Show create teacher form or handle form submission"""
    if request.method == 'POST':
        form = TeacherCreateForm(request.POST)
        if form.is_valid():
            try:
                # Save with request for email URL building
                teacher, generated_password = form.save(request=request)

                # Build success message
                success_message = f'Teacher {teacher.get_full_name()} added successfully!'
                if generated_password:
                    success_message += f' Credentials sent to {teacher.email}'

                # For HTMX requests, return success with OOB swap to refresh list
                if request.headers.get('HX-Request'):
                    # Get updated teacher list with pagination (reset to page 1)
                    teachers = Teacher.objects.select_related('user').filter(is_active=True)
                    paginator = Paginator(teachers, 20)
                    page_obj = paginator.get_page(1)

                    # Count teachers with accounts on current page
                    teachers_with_accounts = sum(1 for teacher in page_obj if teacher.user is not None)

                    context = {
                        'teachers': page_obj,
                        'page_obj': page_obj,
                        'teachers_with_accounts': teachers_with_accounts,
                        'search': '',
                        'status': 'active',
                        'gender': '',
                    }

                    # Render the response with OOB swap for the table and statistics
                    table_html = render_to_string('teachers/partials/teacher_rows.html', context, request=request)
                    stats_html = render_to_string('teachers/partials/teacher_statistics.html', context, request=request)
                    response_html = f'''
                    <div id="teacher-form-container" class="hidden" hx-swap-oob="outerHTML"></div>
                    <div id="teachers-table" hx-swap-oob="innerHTML">{table_html}</div>
                    {stats_html}
                    '''
                    response = HttpResponse(response_html)
                    # Use After-Settle to ensure DOM operations complete before event fires
                    response['HX-Trigger-After-Settle'] = 'teacherCreated'
                    return response

                messages.success(request, success_message)
                return redirect('teachers:list')
            except Exception as e:
                messages.error(request, f'Error creating teacher: {str(e)}')
        # If form is invalid and HTMX, return the form with errors
        elif request.headers.get('HX-Request'):
            return render(request, 'teachers/partials/teacher_form_inline.html', {'form': form, 'action': 'create'})
    else:
        # GET request - show empty form
        form = TeacherCreateForm()

    # For HTMX requests, return only the inline form
    if request.headers.get('HX-Request'):
        response = render(request, 'teachers/partials/teacher_form_inline.html', {'form': form, 'action': 'create'})
        response['HX-Trigger'] = 'formLoaded'
        return response

    # Full page render (fallback)
    return render(request, 'teachers/teacher_create.html', {'form': form})


@school_admin_required
def teacher_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """View teacher details"""
    teacher = get_object_or_404(Teacher, pk=pk)

    # For HTMX requests, return inline detail content
    if request.headers.get('HX-Request'):
        response = render(request, 'teachers/partials/teacher_detail_inline.html', {'teacher': teacher})
        response['HX-Trigger'] = 'formLoaded'
        return response

    return render(request, 'teachers/teacher_detail_page.html', {
        'teacher': teacher
    })


@school_admin_required
def teacher_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit teacher information"""
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == 'POST':
        form = TeacherCreateForm(request.POST, instance=teacher)
        if form.is_valid():
            try:
                # Save changes
                teacher, _ = form.save(request=request)

                # For HTMX requests, return success with OOB swap to refresh list
                if request.headers.get('HX-Request'):
                    # Get updated teacher list with pagination (reset to page 1)
                    teachers = Teacher.objects.select_related('user').filter(is_active=True)
                    paginator = Paginator(teachers, 20)
                    page_obj = paginator.get_page(1)

                    # Count teachers with accounts on current page
                    teachers_with_accounts = sum(1 for teacher in page_obj if teacher.user is not None)

                    context = {
                        'teachers': page_obj,
                        'page_obj': page_obj,
                        'teachers_with_accounts': teachers_with_accounts,
                        'search': '',
                        'status': 'active',
                        'gender': '',
                    }

                    # Render the response with OOB swap for the table and statistics
                    table_html = render_to_string('teachers/partials/teacher_rows.html', context, request=request)
                    stats_html = render_to_string('teachers/partials/teacher_statistics.html', context, request=request)
                    response_html = f'''
                    <div id="teacher-form-container" class="hidden" hx-swap-oob="outerHTML"></div>
                    <div id="teachers-table" hx-swap-oob="innerHTML">{table_html}</div>
                    {stats_html}
                    '''
                    response = HttpResponse(response_html)
                    # Use After-Settle to ensure DOM operations complete before event fires
                    response['HX-Trigger-After-Settle'] = 'teacherUpdated'
                    return response

                messages.success(request, f'Teacher {teacher.get_full_name()} updated successfully!')
                return redirect('teachers:list')
            except Exception as e:
                messages.error(request, f'Error updating teacher: {str(e)}')
        # If form is invalid and HTMX, return the form with errors
        elif request.headers.get('HX-Request'):
            return render(request, 'teachers/partials/teacher_form_inline.html', {
                'form': form,
                'action': 'edit',
                'teacher': teacher
            })
    else:
        # GET request - show form with teacher data
        form = TeacherCreateForm(instance=teacher)

    # For HTMX requests, return only the inline form
    if request.headers.get('HX-Request'):
        response = render(request, 'teachers/partials/teacher_form_inline.html', {
            'form': form,
            'action': 'edit',
            'teacher': teacher
        })
        response['HX-Trigger'] = 'formLoaded'
        return response

    # Full page render (fallback)
    return render(request, 'teachers/teacher_create.html', {
        'form': form,
        'teacher': teacher
    })


@school_admin_required
def teacher_export_pdf(request: HttpRequest, pk: int) -> HttpResponse:
    """Export teacher details to PDF"""
    teacher = get_object_or_404(Teacher, pk=pk)

    # Render HTML template
    html_string = render_to_string('teachers/teacher_pdf.html', {
        'teacher': teacher
    })

    # Generate PDF
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # Create response
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f'teacher_{teacher.employee_id}_{timezone.now().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@school_admin_required
@require_http_methods(["GET", "POST", "DELETE"])
def teacher_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Soft delete teacher or show confirmation modal"""
    teacher = get_object_or_404(Teacher, pk=pk)

    # GET request - show confirmation modal
    if request.method == 'GET' and request.headers.get('HX-Request'):
        return render(request, 'teachers/modals/teacher_delete_content.html', {'teacher': teacher})

    # POST or DELETE request - perform deletion
    if request.method in ['POST', 'DELETE']:
        teacher.is_active = False
        teacher.save()

        # Return empty response to close modal and trigger refresh
        return HttpResponse(headers={'HX-Trigger': 'teacherDeleted'})

    # Fallback redirect
    return redirect('teachers:list')


@school_admin_required
@require_http_methods(["GET", "POST"])
def teacher_bulk_import(request: HttpRequest) -> HttpResponse:
    """
    Bulk import teachers from Excel or CSV file.
    Shows upload form and preview of data to be imported.
    """
    if request.method == 'POST':
        form = TeacherBulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Parse file
                teachers_data = form.parse_file()
                create_accounts = form.cleaned_data['create_user_accounts']

                # Store data in session for processing
                request.session['bulk_import_data'] = teachers_data
                request.session['bulk_import_create_accounts'] = create_accounts

                # Calculate statistics
                valid_count = sum(1 for t in teachers_data if t['valid'])
                invalid_count = len(teachers_data) - valid_count

                context = {
                    'teachers': teachers_data,
                    'total': len(teachers_data),
                    'valid': valid_count,
                    'invalid': invalid_count,
                    'create_accounts': create_accounts,
                }

                # For HTMX requests, return inline preview
                if request.headers.get('HX-Request'):
                    response = render(request, 'teachers/partials/bulk_import_preview_inline.html', context)
                    response['HX-Trigger'] = 'formLoaded'
                    return response

                return render(request, 'teachers/bulk_import_preview.html', context)

            except Exception as e:
                messages.error(request, f'Error parsing file: {str(e)}')
                # For HTMX requests, return form with error
                if request.headers.get('HX-Request'):
                    return render(request, 'teachers/partials/bulk_import_inline.html', {'form': form})
    else:
        form = TeacherBulkImportForm()

    # For HTMX requests, return inline form
    if request.headers.get('HX-Request'):
        response = render(request, 'teachers/partials/bulk_import_inline.html', {'form': form})
        response['HX-Trigger'] = 'formLoaded'
        return response

    return render(request, 'teachers/bulk_import.html', {'form': form})


@school_admin_required
@require_http_methods(["POST"])
def teacher_bulk_import_process(request: HttpRequest) -> HttpResponse:
    """
    Process bulk import after preview confirmation.
    Creates teacher records and optionally user accounts.
    """
    teachers_data = request.session.get('bulk_import_data', [])
    create_accounts = request.session.get('bulk_import_create_accounts', False)

    if not teachers_data:
        messages.error(request, 'No import data found. Please upload a file first.')
        return redirect('teachers:bulk_import')

    success_count = 0
    error_count = 0
    errors = []

    for teacher_data in teachers_data:
        # Skip invalid rows
        if not teacher_data['valid']:
            error_count += 1
            continue

        try:
            # Create teacher
            teacher = Teacher(
                first_name=teacher_data['first_name'],
                last_name=teacher_data['last_name'],
                other_names=teacher_data.get('other_names', ''),
                email=teacher_data['email'],
                gender=teacher_data['gender'],
                phone_number=teacher_data['phone_number'],
                employee_id=teacher_data['employee_id'],
                date_joined=teacher_data['date_joined'],
                date_of_birth=teacher_data.get('date_of_birth'),
            )

            # Create user account if requested (or if per-row flag is set)
            should_create_account = create_accounts or teacher_data.get('create_account', False)
            if should_create_account:
                # Generate password
                password = generate_secure_password()

                # Create user
                user = User.objects.create_teacher(
                    email=teacher_data['email'],
                    password=password
                )
                user.force_password_change = True
                user.save()

                teacher.user = user

            teacher.save()
            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {teacher_data['row_num']}: {str(e)}")

    # Clear session data
    if 'bulk_import_data' in request.session:
        del request.session['bulk_import_data']
    if 'bulk_import_create_accounts' in request.session:
        del request.session['bulk_import_create_accounts']

    # Show results
    success_msg = None
    error_msg = None

    if success_count > 0:
        success_msg = f'Successfully imported {success_count} teacher(s).'
        messages.success(request, success_msg)

    if error_count > 0:
        error_msg = f'Failed to import {error_count} teacher(s).'
        if errors:
            error_msg += ' Errors: ' + '; '.join(errors[:5])  # Show first 5 errors
        messages.error(request, error_msg)

    # For HTMX requests, trigger events with OOB swap
    if request.headers.get('HX-Request'):
        # Get updated teacher list with pagination (reset to page 1)
        teachers = Teacher.objects.select_related('user').filter(is_active=True)
        paginator = Paginator(teachers, 20)
        page_obj = paginator.get_page(1)

        # Count teachers with accounts on current page
        teachers_with_accounts = sum(1 for teacher in page_obj if teacher.user is not None)

        context = {
            'teachers': page_obj,
            'page_obj': page_obj,
            'teachers_with_accounts': teachers_with_accounts,
            'search': '',
            'status': 'active',
            'gender': '',
        }

        # Render the response with OOB swap for the table and statistics
        table_html = render_to_string('teachers/partials/teacher_rows.html', context, request=request)
        stats_html = render_to_string('teachers/partials/teacher_statistics.html', context, request=request)
        response_html = f'''
        <div id="teacher-form-container" class="hidden" hx-swap-oob="outerHTML"></div>
        <div id="teachers-table" hx-swap-oob="innerHTML">{table_html}</div>
        {stats_html}
        '''
        response = HttpResponse(response_html)
        # Use After-Settle to ensure DOM operations complete before event fires
        response['HX-Trigger-After-Settle'] = 'bulkImportCompleted'
        return response

    return redirect('teachers:list')


@school_admin_required
def teacher_download_template(request: HttpRequest) -> HttpResponse:
    """
    Download sample Excel template for bulk import.
    """
    file_format = request.GET.get('format', 'xlsx')

    if file_format == 'csv':
        return _generate_csv_template()
    else:
        return _generate_excel_template()


def _generate_excel_template() -> HttpResponse:
    """Generate Excel template with sample data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teachers"

    # Define headers
    headers = [
        'first_name', 'last_name', 'other_names', 'email',
        'gender', 'phone_number', 'employee_id', 'date_joined',
        'date_of_birth', 'create_account'
    ]

    # Style headers
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Add sample data
    sample_data = [
        [
            'John', 'Doe', 'Michael', 'john.doe@example.com',
            'Male', '+233244123456', 'EMP001', '2024-01-15',
            '1985-05-20', 'yes'
        ],
        [
            'Jane', 'Smith', '', 'jane.smith@example.com',
            'Female', '0244987654', 'EMP002', '2024-02-01',
            '1990-08-15', 'no'
        ],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="teachers_import_template_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response


def _generate_csv_template() -> HttpResponse:
    """Generate CSV template with sample data."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="teachers_import_template_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)

    # Write headers
    headers = [
        'first_name', 'last_name', 'other_names', 'email',
        'gender', 'phone_number', 'employee_id', 'date_joined',
        'date_of_birth', 'create_account'
    ]
    writer.writerow(headers)

    # Write sample data
    sample_data = [
        [
            'John', 'Doe', 'Michael', 'john.doe@example.com',
            'Male', '+233244123456', 'EMP001', '2024-01-15',
            '1985-05-20', 'yes'
        ],
        [
            'Jane', 'Smith', '', 'jane.smith@example.com',
            'Female', '0244987654', 'EMP002', '2024-02-01',
            '1990-08-15', 'no'
        ],
    ]

    for row in sample_data:
        writer.writerow(row)

    return response